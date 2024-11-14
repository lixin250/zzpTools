# python.exe -m pip install --upgrade pip
import pandas as pd  #pip install pandas
import openpyxl  # pip install openpyxl  # 处理excel
from openai import OpenAI  #pip install openai # 新增：导入OpenAI库
import requests  # 确保导入requests库

import nltk  # 确保安装nltk库 # 使用NLP库提取关键词、优点和缺点
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords
from collections import Counter
# 新增：使用jieba进行中文分词  nltk不行，情感和关键词不行
import jieba

# 关键词不行， 没有组合 ，没有正负面，TF-IDF尝试
from sklearn.feature_extraction.text import TfidfVectorizer  # pip install scikit-learn 
from collections import defaultdict

# 关键词不行，导入nltk的词性标注和依存句法分析
from nltk import pos_tag
from nltk.tokenize import word_tokenize
# 使用jieba进行分词和词性标注
import jieba.posseg as pseg  # 导入jieba的词性标注模块

# 下载nltk资源（首次运行时需要）
nltk.download('punkt')
nltk.download('stopwords')

# 新增：创建OpenAI客户端
client = OpenAI(
  api_key='xxxx',  # 添加api_key参数
  organization='org-xxxx',
  project='$proj_xxxx',
)

# 新增：调用ChatGPT API的函数
# sk-proj-xxxx
def analyze_reviews(reviews):
    # 确保传入的reviews参数不为空
    if not reviews:
        print("错误: reviews参数为空")
        return None
    
    api_url = "https://api.openai.com/v1/chat/completions"  # API地址
    headers = {
        "Authorization": f"Bearer sk-proj-xxxxx",  # 替换为你的API密钥
        "Content-Type": "application/json"
    }
    data = {
        "model": "gpt-4o-mini",
        "messages": [
            {"role": "user", "content": f"你现在是一个资深运营，你在一家中国公司工作，公司主要做护眼台灯，2020年开始，基于20多年对家长和孩子需求的洞察，孩视宝进军儿童学习桌椅行业，打造全新产品线，开发儿童学习桌、儿童人体工学椅等儿童家具，以打造更健康的儿童学习光空间。你现在取得了天猫产品的所有评论，做数据分析，针对以下评论：{reviews}。给出产品SKU系列，用户评论 提到的关键词，夸奖的优点、提出的缺点痛点问题，最后给到的优化建议意见"}  # 使用传入的reviews
        ]
    }

    # 新增：发送请求到API
    response = requests.post(api_url, headers=headers, json=data)

    # 新增：检查API响应是否成功
    if response.status_code == 200:
        result = response.json().get('choices')[0].get('message').get('content')
        return result
    else:
        # 新增：输出详细的错误信息以便调试
        print(f"错误: {response.status_code}, {response.text}")  # 输出错误信息
        return None  # 返回None以处理错误情况

# 新增：提取关键词的函数
def extract_keywords(reviews):
    # 使用jieba进行分词
    words = jieba.lcut(reviews)
    
    # 去除停用词
    stop_words = set(stopwords.words('chinese'))  # 停用词列表
    filtered_words = [word for word in words if word not in stop_words and word != 'nan' and len(word) > 1]
    # # 统计关键词出现次数
    word_counts = Counter(filtered_words)

    # 新增：只保留前十名重复的关键词
    top_keywords = dict(word_counts.most_common(20))  # 获取前十名关键词
    return top_keywords    # word_counts 、 top_keywords  返回相关条件 关键词及其出现次数的字典

file_name = 'f:/ZZP/tools/A2BExcel/excel评论分析/分类评论.xlsx'  # 新增：定义文件名变量
# 读取Excel文件
df = pd.read_excel(file_name)

# 新增：创建一个字典以存储相同SKU的评论
sku_reviews = {}
for index, row in df.iterrows():
    sku = row['SKU']
    initial_review = str(row['初评'])
    follow_up_review = str(row['追评'])
    
    # 新增：将评论合并并调用API进行分析
    combined_reviews = initial_review + " " + follow_up_review
    if sku not in sku_reviews:
        sku_reviews[sku] = []
    sku_reviews[sku].append(combined_reviews)

# 新增：分析每个SKU的评论
results = []  # 用于存储每个SKU的分析结果
for sku, reviews in sku_reviews.items():
    # 将所有评论合并为一个字符串
    all_reviews = " ".join(reviews)

    # 新增：提取关键词
    keywords = extract_keywords(all_reviews)
    total_reviews_count = len(reviews)  # 统计评论条数
    print(f"分类产品: {sku} 的评论总数: {total_reviews_count} , 请求分析......")

    analysis_result = analyze_reviews(all_reviews)
    if analysis_result is None:
        print(f"========= 分析失败 SKU: {sku} ================= ")
    else:
        print(f"分类产品: {sku}... 分析完毕 ")
    
    # 将 Counter 转换为 DataFrame
    word_counts_df = pd.DataFrame(keywords.items(), columns=['Keyword', 'Count'])
    # 新增：按次数排序关键词
    sorted_keywords = dict(sorted(keywords.items(), key=lambda item: item[1], reverse=True))  # 按次数降序排序
    # 新增：将关键词统计转换为字符串
    keywords_str = ', '.join([f"{k}: {v}" for k, v in sorted_keywords.items()])  # 转换为字符串格式


    # 新增：将分析结果添加到结果列表
    results.append({
        'SKU': sku,
        'result': analysis_result,
        'all_reviews_content': all_reviews,

        'keywords': keywords_str,  # 添加关键词统计
        'total_reviews_count': total_reviews_count,  # 添加评论总数
    })

wb = openpyxl.load_workbook(file_name)  # 加载现有的Excel文件
# 检查工作表是否存在
if '分析结果' in wb.sheetnames:
    ws_results = wb['分析结果']  # 获取已有的工作表
    ws_results.delete_rows(1, ws_results.max_row)  # 清空原有内容，从第二行开始删除
else:
    ws_results = wb.create_sheet(title="分析结果")  # 创建新工作表

ws_results.append(['SKU', "统计关键词", "评论数", "结论", "分析内容包括"])  # 添加表头
    
for result in results:
    ws_results.append([
        result['SKU'],
        result['keywords'],
        result['total_reviews_count'],
        result['result'],
        result['all_reviews_content']
    ])

# 保存Excel文件
try:
    wb.save(file_name)  # 尝试保存文件
except PermissionError as e:
    print(f"权限错误: {e}. 请确保文件没有被其他程序占用或没有写入权限。")  # 输出错误信息
print("完毕~~~~~ ")
